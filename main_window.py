"""
Главное окно приложения Word-шаблонов.
Интерфейс загружается из .ui файла в рантайме (QUiLoader), без генерации в Python.
"""
import logging
import sys
from pathlib import Path

from PySide6.QtCore import QBuffer, QByteArray, QFile, QIODevice, Qt, QObject, Signal, QEvent
from PySide6.QtCore import QSettings
from PySide6.QtGui import QAction
from PySide6.QtWidgets import (
    QComboBox,
    QDialog,
    QFrame,
    QFileDialog,
    QHBoxLayout,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPlainTextEdit,
    QPushButton,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from db.connection import Connection
from db.models import DocumentTemplate
from logger import py_logger
import settings as st
from services.template_service import (
    get_template_linked_data_table,
    get_template_loop_blocks,
    load_templates_list,
    register_template,
    get_template_fields,
)
from services.field_value_service import get_values_by_template_id
from services.generation_service import generate_document
from services.data_table_service import list_records, get_record_values
from settings_dialog import SettingsDialog, get_settings_interface_log_expanded, get_settings_interface_templates_dir
from excel_import_dialog import ExcelImportDialog
from template_data_binding_dialog import TemplateDataBindingDialog


class LogSignalBridge(QObject):
    """Мост для передачи сообщений лога в UI из любого потока (сигнал очередится в главный поток)."""
    log_message = Signal(str)


class QtLogHandler(logging.Handler):
    """Обработчик логирования, перенаправляющий сообщения в Qt-сигнал."""
    def __init__(self, signal_bridge: LogSignalBridge):
        super().__init__()
        self._bridge = signal_bridge

    def emit(self, record: logging.LogRecord):
        try:
            msg = self.format(record)
            self._bridge.log_message.emit(msg)
        except Exception:
            self.handleError(record)


class MainWindow(QMainWindow):
    _tables_combobox = None
    _data_tables = None
    _current_fields = None  # список (field_name, display_name) для строк таблицы
    _load_excel_btn = None
    _record_combobox = None
    _record_choice_widget = None  # виджет «Запись из набора данных», показывать только при linked_data_table
    _loop_tab_widgets = None
    _loop_blocks_meta = None

    def __init__(self, ui_file_name: str = "ui/main_window.ui", ui_loader=None):
        super().__init__()
        self._tables_combobox = None
        self._data_tables = None
        self._current_fields = []
        self._load_excel_btn = None
        self._loop_tab_widgets = {}
        self._loop_blocks_meta = {}
        loader = ui_loader if ui_loader is not None else QUiLoader()

        ui_path = Path(ui_file_name)
        if not ui_path.is_absolute():
            base = Path(__file__).resolve().parent
            ui_path = (base / ui_file_name).resolve()
        if not ui_path.exists():
            py_logger.warning("UI file not found: %s", ui_path)
            self.window = QWidget(self)
            self.setCentralWidget(self.window)
            self.setWindowTitle("Word-шаблоны")
            self._setup_stub_ui()
            self.connection = Connection()
            return

        ui_file = QFile(str(ui_path))
        if ui_file.open(QIODevice.ReadOnly):
            device = ui_file
        else:
            ui_file.close()
            try:
                data = ui_path.read_bytes()
            except OSError as e:
                py_logger.error("Cannot read UI file %s: %s", ui_path, e)
                sys.exit(-1)
            device = QBuffer(QByteArray(data))
            if not device.open(QIODevice.ReadOnly):
                py_logger.error("Cannot open UI buffer")
                sys.exit(-1)

        loaded = loader.load(device, None)
        if hasattr(device, "close"):
            device.close()

        if not loaded:
            py_logger.error("QUiLoader: %s", loader.errorString())
            sys.exit(-1)

        # Ссылку храним до изъятия виджетов, иначе C++ объект может быть удалён
        self._loaded_ui = loaded
        cw = loaded.centralWidget()
        mb = loaded.menuBar()
        sb = loaded.statusBar()
        self.setCentralWidget(cw)
        if mb is not None:
            self.setMenuBar(mb)
        if sb is not None:
            self.setStatusBar(sb)
        self.window = self.centralWidget()
        self.connection = Connection()
        self._bind_ui()
        self._setup_log_panel()
        self._setup_menu_settings()
        self.setWindowTitle("Word-шаблоны")
        self._restore_geometry_state()
        py_logger.info("MainWindow initialized (Word templates)")

    def _setup_stub_ui(self):
        """Минимальный UI, если .ui файл не найден."""
        from PySide6.QtWidgets import QLabel, QVBoxLayout
        layout = QVBoxLayout(self.window)
        layout.addWidget(QLabel("Файл ui/main_window.ui не найден.\nШаблоны: см. БД document_templates."))

    def _find(self, type_, name):
        """Виджет по имени в центральном виджете."""
        if self.window is None:
            return None
        return self.window.findChild(type_, name)

    def _bind_ui(self):
        """Привязка виджетов из загруженного .ui к логике."""
        self._tables_combobox = self._find(QComboBox, "tables_combobox")
        if self._tables_combobox is not None:
            self._refresh_templates_combo()
            self._tables_combobox.currentIndexChanged.connect(self._on_template_changed)

        add_file_btn = self._find(QPushButton, "add_file_pushbutton")
        if add_file_btn is not None:
            add_file_btn.clicked.connect(self._on_add_template)

        export_btn = self._find(QPushButton, "export_pushbutton")
        if export_btn is not None:
            export_btn.clicked.connect(self._on_generate)

        self._data_tables = self._find(QTableWidget, "data_tables")
        if self._data_tables is not None:
            self._data_tables.horizontalHeader().setStretchLastSection(True)

        self._load_excel_btn = self._find(QPushButton, "load_excel_pushbutton")
        if self._load_excel_btn is not None:
            self._load_excel_btn.clicked.connect(self._on_load_from_excel)

        self._record_combobox = self._find(QComboBox, "record_combobox")
        self._record_choice_widget = self._find(QWidget, "record_choice_widget")
        if self._record_choice_widget is not None:
            self._record_choice_widget.setVisible(False)
        if self._record_combobox is not None:
            self._record_combobox.currentIndexChanged.connect(self._on_record_selected)
        self._wrap_fields_in_tabs()

    def _setup_log_panel(self):
        """Скрываемая/раскрываемая панель лога снизу главного окна."""
        layout = self.window.layout()
        if layout is None:
            return

        self._log_bridge = LogSignalBridge(self)
        self._log_edit = QPlainTextEdit(self.window)
        self._log_edit.setReadOnly(True)
        self._log_edit.setMinimumHeight(80)
        self._log_edit.setMaximumHeight(200)
        self._log_edit.setMaximumBlockCount(2000)
        self._log_edit.setPlaceholderText("Лог (раскройте кнопкой «Лог ▼»)")
        self._log_edit.hide()

        self._log_toggle_btn = QPushButton("Лог ▼", self.window)
        self._log_toggle_btn.setToolTip("Показать/скрыть панель лога")
        self._log_toggle_btn.setMaximumWidth(80)
        self._log_clear_btn = QPushButton("Очистить", self.window)
        self._log_clear_btn.setMaximumWidth(80)

        bar = QWidget(self.window)
        bar_layout = QHBoxLayout(bar)
        bar_layout.setContentsMargins(0, 4, 0, 0)
        bar_layout.addWidget(self._log_toggle_btn)
        bar_layout.addWidget(self._log_clear_btn)
        bar_layout.addStretch()

        log_container = QFrame(self.window)
        log_container.setFrameStyle(QFrame.Shape.StyledPanel)
        log_v = QVBoxLayout(log_container)
        log_v.setContentsMargins(4, 0, 4, 4)
        log_v.addWidget(bar)
        log_v.addWidget(self._log_edit)

        self._log_toggle_btn.clicked.connect(self._on_log_toggle)
        self._log_clear_btn.clicked.connect(lambda: self._log_edit.clear())
        self._log_bridge.log_message.connect(self._append_log_message)

        handler = QtLogHandler(self._log_bridge)
        handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        py_logger.addHandler(handler)
        self._qt_log_handler = handler

        layout.addWidget(log_container)

    def _on_log_toggle(self):
        if self._log_edit.isVisible():
            self._log_edit.hide()
            self._log_toggle_btn.setText("Лог ▼")
        else:
            self._log_edit.show()
            self._log_toggle_btn.setText("Лог ▲")

    def _append_log_message(self, msg: str):
        self._log_edit.appendPlainText(msg)
        self._log_edit.verticalScrollBar().setValue(
            self._log_edit.verticalScrollBar().maximum()
        )

    def _settings_obj(self) -> QSettings:
        return QSettings()

    def _restore_geometry_state(self):
        """Восстановить размер, позицию и состояние окна из QSettings; иначе — на весь экран."""
        s = self._settings_obj()
        geom = s.value("MainWindow/geometry")
        state = s.value("MainWindow/state")
        if isinstance(geom, QByteArray) and geom:
            self.restoreGeometry(geom)
        if isinstance(state, QByteArray) and state:
            self.restoreState(state)
        if not (geom and state):
            self.setWindowState(Qt.WindowState.WindowMaximized)
        if getattr(self, "_log_edit", None) and get_settings_interface_log_expanded():
            if self._log_edit.isHidden():
                self._on_log_toggle()

    def _save_geometry_state(self):
        """Сохранить геометрию и состояние окна в QSettings."""
        s = self._settings_obj()
        s.setValue("MainWindow/geometry", self.saveGeometry())
        s.setValue("MainWindow/state", self.saveState())
        s.sync()

    def closeEvent(self, event):
        self._save_geometry_state()
        super().closeEvent(event)

    def changeEvent(self, event):
        if event.type() == QEvent.Type.WindowStateChange:
            self._save_geometry_state()
        super().changeEvent(event)

    def _setup_menu_settings(self):
        """Добавить в меню бар «Сервис» (Настройки, Импорт в БД) и «Данные»."""
        menubar = self.menuBar()
        if menubar is None:
            return
        menu_serv = menubar.addMenu("Сервис")
        act = QAction("Настройки…", self)
        act.triggered.connect(self._on_settings)
        menu_serv.addAction(act)
        act_import = QAction("Импорт из Excel в БД…", self)
        act_import.triggered.connect(self._on_import_excel_to_db)
        menu_serv.addAction(act_import)
        act_bind = QAction("Привязать шаблон к набору данных…", self)
        act_bind.triggered.connect(self._on_bind_template_to_data)
        menu_serv.addAction(act_bind)

    def _on_import_excel_to_db(self):
        """Открыть диалог импорта Excel в набор данных."""
        dlg = ExcelImportDialog(self)
        dlg.exec()

    def _on_bind_template_to_data(self):
        """Открыть диалог привязки шаблона к набору данных."""
        template_id = self._tables_combobox.currentData() if self._tables_combobox else None
        dlg = TemplateDataBindingDialog(self, initial_template_id=template_id)
        dlg.exec()
        # Обновить отображение выбора записи, если выбран тот же шаблон
        if self._tables_combobox and self._tables_combobox.currentIndex() > 0:
            self._on_template_changed(self._tables_combobox.currentIndex())

    def _on_settings(self):
        """Открыть модальное окно настроек."""
        dlg = SettingsDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            st.DEFAULT_TEMPLATES_DIR = get_settings_interface_templates_dir()

    def _on_load_from_excel(self):
        """Загрузить значения из Excel (колонки: имя поля, значение) и подставить в таблицу."""
        if not self._current_fields:
            QMessageBox.warning(
                self,
                "Загрузка из Excel",
                "Сначала выберите шаблон.",
            )
            return
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл Excel",
            "",
            "Excel (*.xlsx *.xls);;All (*.*)",
        )
        if not path:
            return
        try:
            data = self._read_name_value_excel(path)
        except Exception as e:
            py_logger.error("Excel read: %s", e)
            QMessageBox.critical(
                self,
                "Ошибка",
                f"Не удалось прочитать файл:\n{e}",
            )
            return
        if not data:
            QMessageBox.warning(
                self,
                "Загрузка из Excel",
                "В файле не найдено колонок «имя» и «значение» (или «name» и «value»).",
            )
            return
        py_logger.info(
            "---------- Вставка полей из Excel (простые поля) ---------- Файл: %s, прочитано пар имя/значение: %s",
            path, len(data),
        )
        for k, v in data.items():
            py_logger.info("  Excel поле «%s» = %r", k, v)
        self._apply_values_to_table(data, source_label=f"Excel {path}")
        QMessageBox.information(
            self,
            "Загрузка из Excel",
            f"Подставлено значений: {len(data)}.",
        )

    def _cell_to_str(self, c) -> str:
        """Ячейка в строку (числа, даты, None)."""
        if c is None:
            return ""
        if hasattr(c, "isoformat"):
            return c.isoformat()
        if isinstance(c, (int, float)):
            return str(int(c)) if isinstance(c, float) and c == int(c) else str(c)
        return str(c).strip()

    def _read_name_value_excel(self, path: str) -> dict:
        """
        Прочитать Excel: две колонки — имя поля и значение.
        Поддерживаются заголовки: имя/name/поле, значение/value.
        Возвращает словарь {имя_поля: значение} (ключи в верхнем регистре).
        """
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        if ws is None:
            return {}
        rows = [[self._cell_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
        if not rows:
            return {}
        first = [c.lower().strip() for c in rows[0] if c]
        name_col = value_col = None
        for col, header in enumerate(rows[0]):
            h = (header or "").strip().lower()
            if h in ("имя", "name", "поле", "field"):
                name_col = col
            elif h in ("значение", "value"):
                value_col = col
        # Требуем обе колонки: «имя» и «значение». Иначе файл для другого формата (например цикл).
        if name_col is None or value_col is None:
            return {}
        data = {}
        for row in rows[1:] if len(rows) > 1 else []:
            if not row or name_col >= len(row):
                continue
            name = self._cell_to_str(row[name_col]).strip()
            if not name:
                continue
            value = self._cell_to_str(row[value_col]) if value_col < len(row) else ""
            data[name.upper()] = value
        return data

    def _apply_values_to_table(self, data: dict, source_label: str = "источник"):
        """
        Подставить в таблицу полей только те значения из data, для которых есть ключ
        (имя_поля -> значение). Поля, которых нет в источнике (другой Excel, запись),
        не изменяются — данные из разных Excel можно подставлять по частям.
        """
        if self._data_tables is None or not self._current_fields:
            return
        py_logger.info(
            "Заполнение полей (применение значений): источник=%s, полей в данных: %s, ключи: %s",
            source_label, len(data), list(data.keys()),
        )
        data_upper = {k.upper(): v for k, v in data.items()}
        applied = 0
        for row in range(self._data_tables.rowCount()):
            if row >= len(self._current_fields):
                break
            field_name, _ = self._current_fields[row]
            key = field_name.upper()
            if key not in data_upper and field_name not in data:
                continue  # поля нет в источнике — не трогаем ячейку
            value = data_upper.get(key, data.get(field_name, ""))
            w = self._data_tables.cellWidget(row, 1)
            if w is None:
                continue
            if isinstance(w, QComboBox):
                idx = w.findData(value)
                if idx >= 0:
                    w.setCurrentIndex(idx)
                    applied += 1
                else:
                    idx = w.findText(value)
                    if idx >= 0:
                        w.setCurrentIndex(idx)
                        applied += 1
                    else:
                        w.setCurrentIndex(0)
            elif isinstance(w, QLineEdit):
                w.setText(value)
                applied += 1
        py_logger.info(
            "Заполнение полей: применено к ячейкам: %s из %s полей таблицы; значения: %s",
            applied, len(self._current_fields),
            {fn: data_upper.get(fn.upper(), data.get(fn, "")) for fn, _ in self._current_fields if fn.upper() in data_upper or fn in data},
        )

    def _refresh_templates_combo(self):
        if self._tables_combobox is None:
            return
        self._tables_combobox.clear()
        self._tables_combobox.addItem(st.AUX.NOT_CHOOSED_ITEM)
        try:
            for t in load_templates_list():
                self._tables_combobox.addItem(t.name, t.id)
        except Exception as e:
            py_logger.error("Load templates: %s", e)

    def _on_template_changed(self, index: int):
        if index <= 0:
            self._current_fields = []
            if self._data_tables is not None:
                self._data_tables.setRowCount(0)
            if self._load_excel_btn is not None:
                self._load_excel_btn.setEnabled(False)
            self._hide_record_choice()
            self._rebuild_loop_tabs(None)
            return
        if self._tables_combobox is None:
            return
        template_id = self._tables_combobox.currentData()
        if template_id is None:
            return
        py_logger.info("Template selected: id=%s", template_id)
        self._fill_fields_table(template_id)
        self._rebuild_loop_tabs(template_id)
        if self._load_excel_btn is not None:
            self._load_excel_btn.setEnabled(True)
        linked = get_template_linked_data_table(template_id)
        if linked:
            self._show_record_choice(linked.id)
        else:
            self._hide_record_choice()

    def _hide_record_choice(self):
        """Скрыть блок выбора записи и очистить комбобокс."""
        if self._record_choice_widget is not None:
            self._record_choice_widget.setVisible(False)
        if self._record_combobox is not None:
            self._record_combobox.blockSignals(True)
            self._record_combobox.clear()
            self._record_combobox.addItem(st.AUX.NOT_CHOOSED_ITEM, None)
            self._record_combobox.blockSignals(False)

    def _show_record_choice(self, data_table_id: int):
        """Показать блок выбора записи и заполнить комбобокс записями набора данных."""
        if self._record_choice_widget is None or self._record_combobox is None:
            return
        records = list_records(data_table_id)
        self._record_combobox.blockSignals(True)
        self._record_combobox.clear()
        self._record_combobox.addItem(st.AUX.NOT_CHOOSED_ITEM, None)
        for rec in records:
            vals = get_record_values(rec.id)
            # Краткая подпись: первые два значения или «Запись N»
            parts = list(vals.values())[:2]
            label = " | ".join(str(p) for p in parts if p) or f"Запись {rec.id}"
            if len(label) > 50:
                label = label[:47] + "..."
            self._record_combobox.addItem(label, rec.id)
        self._record_combobox.blockSignals(False)
        self._record_choice_widget.setVisible(True)

    def _on_record_selected(self, index: int):
        """Подставить в таблицу полей значения выбранной записи из набора данных."""
        if index <= 0 or self._record_combobox is None:
            return
        record_id = self._record_combobox.currentData()
        if record_id is None:
            return
        try:
            values = get_record_values(record_id)
        except Exception as e:
            py_logger.error("Record values: %s", e)
            return
        py_logger.info(
            "---------- Заполнение полей из записи ---------- record_id=%s, полей в записи: %s",
            record_id, len(values),
        )
        for k, v in values.items():
            py_logger.info("  Запись поле «%s» = %r", k, v)
        # Привести к формату имя_поля -> значение для _apply_values_to_table (ожидает ключи как в шаблоне)
        self._apply_values_to_table(values, source_label=f"запись id={record_id}")

    def _fill_fields_table(self, template_id: int):
        """Заполнить таблицу полей и виджетами ввода/выбора значения."""
        if self._data_tables is None:
            return
        fields = get_template_fields(template_id)
        values_by_field = get_values_by_template_id(template_id)
        self._current_fields = [(f.field_name, f.display_name or f.field_name) for f in fields]
        self._data_tables.setRowCount(len(fields))
        for row, f in enumerate(fields):
            name_item = QTableWidgetItem(f.display_name or f.field_name)
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            self._data_tables.setItem(row, 0, name_item)
            field_values = values_by_field.get(f.id, [])
            if field_values:
                combo = QComboBox()
                combo.addItem(st.AUX.NOT_CHOOSED_ITEM, None)
                for v in field_values:
                    combo.addItem(v.value_text, v.value_text)
                default = next((v for v in field_values if v.is_default), None)
                if default:
                    combo.setCurrentIndex(combo.findData(default.value_text))
                else:
                    combo.setCurrentIndex(0)
                self._data_tables.setCellWidget(row, 1, combo)
            else:
                line = QLineEdit()
                if f.default_value:
                    line.setPlaceholderText(f.default_value)
                self._data_tables.setCellWidget(row, 1, line)
        py_logger.info(
            "---------- Заполнение полей (таблица) ---------- шаблон id=%s, отображено полей: %s, имена: %s",
            template_id, len(fields), [f.field_name for f in fields],
        )

    def _wrap_fields_in_tabs(self):
        if self._data_tables is None:
            return
        parent = self._data_tables.parent()
        if parent is None:
            return
        layout = parent.layout()
        if layout is None:
            return
        idx = layout.indexOf(self._data_tables)
        if idx < 0:
            return
        layout.removeWidget(self._data_tables)
        self._fields_tab_widget = QTabWidget(parent)
        simple_tab = QWidget()
        tab_layout = QVBoxLayout(simple_tab)
        tab_layout.setContentsMargins(0, 0, 0, 0)
        tab_layout.addWidget(self._data_tables)
        self._fields_tab_widget.addTab(simple_tab, "Поля")
        layout.insertWidget(idx, self._fields_tab_widget)

    def _rebuild_loop_tabs(self, template_id):
        if not hasattr(self, "_fields_tab_widget"):
            return
        while self._fields_tab_widget.count() > 1:
            self._fields_tab_widget.removeTab(1)
        self._loop_tab_widgets.clear()
        self._loop_blocks_meta.clear()
        if template_id is None:
            return
        try:
            blocks = get_template_loop_blocks(template_id)
        except Exception as e:
            py_logger.error("Loop blocks load: %s", e)
            return
        for block in blocks:
            self._loop_blocks_meta[block.id] = block
            tab_widget, tbl = self._create_loop_tab_widget(block)
            self._fields_tab_widget.addTab(tab_widget, block.label or block.loop_var)
            self._loop_tab_widgets[block.id] = tbl

    def _create_loop_tab_widget(self, block):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(4, 4, 4, 4)
        btn_bar = QWidget()
        btn_layout = QHBoxLayout(btn_bar)
        btn_layout.setContentsMargins(0, 0, 0, 0)
        add_btn = QPushButton("+ Строка")
        del_btn = QPushButton("− Строка")
        excel_btn = QPushButton("Загрузить из Excel…")
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(del_btn)
        btn_layout.addWidget(excel_btn)
        btn_layout.addStretch()
        display_names = [lf.display_name or lf.field_name for lf in block.loop_fields]
        tbl = QTableWidget(0, len(display_names))
        tbl.setHorizontalHeaderLabels(display_names)
        tbl.horizontalHeader().setStretchLastSection(True)
        col_count = len(display_names)
        add_btn.clicked.connect(lambda: self._loop_add_row(tbl, col_count))
        del_btn.clicked.connect(lambda: self._loop_del_row(tbl))
        excel_btn.clicked.connect(
            lambda checked=False, b=block, t=tbl: self._loop_load_excel(b, t)
        )
        layout.addWidget(btn_bar)
        layout.addWidget(tbl)
        return tab, tbl

    def _loop_add_row(self, tbl: QTableWidget, col_count: int):
        r = tbl.rowCount()
        tbl.insertRow(r)
        for c in range(col_count):
            tbl.setItem(r, c, QTableWidgetItem(""))

    def _loop_del_row(self, tbl: QTableWidget):
        row = tbl.currentRow()
        if row < 0:
            row = tbl.rowCount() - 1
        if row >= 0:
            tbl.removeRow(row)

    def _loop_load_excel(self, block, tbl: QTableWidget):
        path, _ = QFileDialog.getOpenFileName(
            self,
            f"Excel для цикла «{block.label or block.loop_var}»",
            "",
            "Excel (*.xlsx *.xls);;All (*.*)",
        )
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            if ws is None:
                QMessageBox.warning(self, "Загрузка Excel", "Файл пуст.")
                return
            raw_rows = [
                [self._cell_to_str(c) for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
            if not raw_rows:
                QMessageBox.warning(self, "Загрузка Excel", "Файл пуст.")
                return

            def _norm(h: str) -> str:
                h = h.strip()
                if h.startswith("{{") and h.endswith("}}"):
                    h = h[2:-2].strip()
                return h.upper()

            norm_headers = [_norm(h) for h in raw_rows[0]]
            block_fields = [lf.field_name for lf in block.loop_fields]
            block_fields_upper = [f.upper() for f in block_fields]
            col_map = {
                xi: block_fields_upper.index(nh)
                for xi, nh in enumerate(norm_headers)
                if nh in block_fields_upper
            }
            py_logger.info(
                "---------- Вставка полей из Excel (цикл) ---------- Цикл: %s, файл: %s",
                block.loop_var, path,
            )
            py_logger.info(
                "  Заголовки в файле: %s; поля цикла: %s; сопоставлено колонок: %s",
                raw_rows[0], block_fields, dict(col_map),
            )
            tbl.setRowCount(0)
            for row_vals in raw_rows[1:]:
                if all(self._cell_to_str(v) == "" for v in row_vals):
                    continue
                r = tbl.rowCount()
                tbl.insertRow(r)
                for xi, ti in col_map.items():
                    val = self._cell_to_str(row_vals[xi]) if xi < len(row_vals) else ""
                    tbl.setItem(r, ti, QTableWidgetItem(val))
            py_logger.info("  Загружено строк в таблицу цикла: %s", tbl.rowCount())
            for r in range(min(5, tbl.rowCount())):
                row_data = {
                    block_fields[c]: (tbl.item(r, c).text() if tbl.item(r, c) else "")
                    for c in range(len(block_fields))
                }
                py_logger.info("  Строка %s: %s", r + 1, row_data)
            if tbl.rowCount() > 5:
                py_logger.info("  ... и ещё %s строк", tbl.rowCount() - 5)
            QMessageBox.information(
                self, "Загрузка Excel", f"Загружено строк: {tbl.rowCount()}."
            )
        except Exception as e:
            py_logger.error("Loop Excel load: %s", e)
            QMessageBox.critical(self, "Ошибка", f"Не удалось прочитать файл:\n{e}")

    def _on_add_template(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл шаблона .docx",
            st.DEFAULT_TEMPLATES_DIR or "",
            "Word (*.docx);;All (*.*)",
        )
        if not path:
            return
        path = Path(path)
        name = path.stem.replace("_", " ").strip() or "Шаблон"
        code = path.stem.replace(" ", "_").strip() or "template"
        template, errors = register_template(
            name=name,
            code=code,
            file_path=str(path),
            sync_fields_from_file=True,
        )
        if errors:
            QMessageBox.warning(self, "Ошибка загрузки шаблона", "\n".join(errors))
            return
        if template:
            self._refresh_templates_combo()
            QMessageBox.information(
                self,
                "Шаблон загружен",
                f"Добавлен шаблон «{template.name}» с полями из файла.",
            )

    def _on_generate(self):
        if self._tables_combobox is None:
            return
        template_id = self._tables_combobox.currentData()
        if template_id is None or self._tables_combobox.currentIndex() <= 0:
            QMessageBox.warning(
                self,
                "Генерация",
                "Выберите шаблон.",
            )
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить документ",
            "",
            "Word (*.docx);;All (*.*)",
        )
        if not path:
            return
        if not path.lower().endswith(".docx"):
            path = path + ".docx"
        context = self._collect_context_from_table()
        py_logger.info(
            "Generate: template_id=%s, path=%s, context_keys=%s",
            template_id, path, list(context.keys()),
        )
        try:
            generate_document(template_id, context, path)
            QMessageBox.information(
                self,
                "Генерация",
                f"Документ сохранён:\n{path}",
            )
        except Exception as e:
            py_logger.error("Generate: %s", e)
            QMessageBox.critical(
                self,
                "Ошибка",
                str(e),
            )

    def _collect_context_from_table(self) -> dict:
        """Собрать из таблицы полей словарь field_name -> value для подстановки в шаблон."""
        context = {}
        if self._data_tables is not None and self._current_fields:
            for row in range(self._data_tables.rowCount()):
                if row >= len(self._current_fields):
                    break
                field_name, _ = self._current_fields[row]
                w = self._data_tables.cellWidget(row, 1)
                val = ""
                if w is not None:
                    if isinstance(w, QComboBox):
                        v = w.currentData()
                        if v is not None:
                            val = v if isinstance(v, str) else str(v)
                    elif isinstance(w, QLineEdit):
                        val = w.text().strip()
                context[field_name] = val
                py_logger.debug(
                    "Context field: row=%s, field_name=%s, widget=%s, value=%r",
                    row, field_name, type(w).__name__ if w else None, val,
                )
        for block_id, tbl in self._loop_tab_widgets.items():
            block = self._loop_blocks_meta.get(block_id)
            if block is None:
                continue
            fields = [lf.field_name for lf in block.loop_fields]
            rows = []
            for r in range(tbl.rowCount()):
                row_dict = {
                    field: (tbl.item(r, c).text() if tbl.item(r, c) else "")
                    for c, field in enumerate(fields)
                }
                if any(v.strip() for v in row_dict.values()):
                    rows.append(row_dict)
            context[block.loop_var] = rows
            py_logger.debug(
                "Context loop: loop_var=%s, rows_count=%s, fields=%s, sample_row=%s",
                block.loop_var, len(rows), fields, rows[0] if rows else None,
            )
        py_logger.info(
            "Context collected: %s simple keys, %s loop keys; simple_values=%s",
            len([k for k in context if not isinstance(context.get(k), list)]),
            len([k for k in context if isinstance(context.get(k), list)]),
            {k: v for k, v in context.items() if not isinstance(v, list)},
        )
        return context
