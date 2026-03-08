"""
Создание тестовых Excel-файлов (имя поля / значение) для загрузки в приложение.
Формат: колонки «имя» и «значение» — для режима «Из Excel».

Запуск из корня проекта:
  python -m tests.create_sample_excel
  python tests/create_sample_excel.py
"""
import sys
from pathlib import Path

if __name__ == "__main__" and __package__ is None:
    _root = Path(__file__).resolve().parent.parent
    sys.path.insert(0, str(_root))

import openpyxl
from openpyxl.styles import Font


def create_excel_file(path: Path, headers: tuple, rows: list) -> None:
    """Записать в path лист с заголовками и строками."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        return
    ws.title = "Данные"
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    wb.save(str(path))
    print(f"Создан файл: {path}")


def main():
    root = Path(__file__).resolve().parent
    out_dir = root / "fixtures"

    # Договор: поля из contract_services (формат имя/значение для «Вставить из Excel»)
    create_excel_file(
        out_dir / "contract_services_values.xlsx",
        ("имя", "значение"),
        [
            ("CLIENT_NAME", "ООО Рога и копыта"),
            ("CONTRACT_NUMBER", "1"),
            ("CONTRACT_DATE", "2025-01-15"),
            ("CONTRACT_TYPE", "Разовый"),
            ("AMOUNT", "15000"),
        ],
    )

    # Акт: поля из act_works (формат имя/значение для «Вставить из Excel»)
    create_excel_file(
        out_dir / "act_works_values.xlsx",
        ("имя", "значение"),
        [
            ("ACT_NUMBER", "АКТ-001"),
            ("ACT_DATE", "2025-02-01"),
            ("CLIENT_NAME", "ООО Рога и копыта"),
            ("EXECUTOR_NAME", "ИП Исполнитель"),
            ("WORK_DESCRIPTION", "Консультационные услуги за январь 2025"),
        ],
    )

    # Счёт на оплату (простой): поля INVOICE_NO, DATE, CUSTOMER, TOTAL
    create_excel_file(
        out_dir / "invoice_simple_values.xlsx",
        ("имя", "значение"),
        [
            ("INVOICE_NO", "СЧ-2025-001"),
            ("DATE", "2025-03-10"),
            ("CUSTOMER", "ООО Покупатель"),
            ("TOTAL", "75000"),
        ],
    )

    # Счёт-фактура: простые поля (формат имя/значение)
    create_excel_file(
        out_dir / "invoice_loop_values.xlsx",
        ("имя", "значение"),
        [
            ("INVOICE_NUMBER", "СФ-2025-001"),
            ("INVOICE_DATE", "2025-03-01"),
            ("CLIENT_NAME", "ООО Рога и копыта"),
        ],
    )

    # Счёт-фактура: данные цикла items — заголовки {{FIELD}} (тест нотации с фигурными скобками)
    create_excel_file(
        out_dir / "invoice_loop_items.xlsx",
        ("{{NAME}}", "{{QTY}}", "{{PRICE}}"),
        [
            ("Разработка модуля авторизации", "1", "45000"),
            ("Техническая поддержка", "8", "3500"),
            ("Тестирование и документация", "1", "15000"),
        ],
    )

    # Акт сдачи-приёмки: простые поля (формат имя/значение)
    create_excel_file(
        out_dir / "mixed_loop_values.xlsx",
        ("имя", "значение"),
        [
            ("EXECUTOR", "ООО СофтДев"),
            ("CONTRACT_DATE", "2025-03-15"),
        ],
    )

    # Акт сдачи-приёмки: данные цикла works — заголовки без фигурных скобок
    create_excel_file(
        out_dir / "mixed_loop_works.xlsx",
        ("WORK", "AMOUNT"),
        [
            ("Анализ требований", "20000"),
            ("Проектирование архитектуры", "35000"),
            ("Разработка MVP", "80000"),
            ("Интеграционное тестирование", "15000"),
        ],
    )

    # Накладная: простые поля (формат имя/значение)
    create_excel_file(
        out_dir / "delivery_note_values.xlsx",
        ("имя", "значение"),
        [
            ("CONSIGNEE", "ООО Складской комплекс"),
            ("DELIVERY_DATE", "2025-03-20"),
        ],
    )

    # Накладная: данные цикла goods (NAME, QTY, UNIT)
    create_excel_file(
        out_dir / "delivery_note_goods.xlsx",
        ("NAME", "QTY", "UNIT"),
        [
            ("Ноутбук Dell XPS 15", "2", "шт."),
            ("Мышь беспроводная", "5", "шт."),
            ("Клавиатура механическая", "3", "шт."),
            ("Монитор 27\"", "2", "шт."),
        ],
    )

    # Отчёт по проекту (3 цикла): простые поля
    create_excel_file(
        out_dir / "project_report_values.xlsx",
        ("имя", "значение"),
        [
            ("PROJECT_NAME", "Внедрение CRM"),
            ("REPORT_DATE", "2025-04-01"),
            ("AUTHOR", "Иванов И.И."),
        ],
    )
    # Цикл 1: участники (NAME, ROLE, CONTACT)
    create_excel_file(
        out_dir / "project_report_participants.xlsx",
        ("NAME", "ROLE", "CONTACT"),
        [
            ("Петров П.П.", "Руководитель", "petrov@company.ru"),
            ("Сидорова С.С.", "Аналитик", "sidorova@company.ru"),
            ("Козлов К.К.", "Разработчик", "kozlov@company.ru"),
        ],
    )
    # Цикл 2: этапы (MILESTONE, DEADLINE, STATUS)
    create_excel_file(
        out_dir / "project_report_milestones.xlsx",
        ("MILESTONE", "DEADLINE", "STATUS"),
        [
            ("Анализ требований", "2025-02-15", "Выполнено"),
            ("Проектирование", "2025-03-01", "Выполнено"),
            ("Разработка", "2025-04-15", "В работе"),
            ("Тестирование", "2025-05-01", "Запланировано"),
        ],
    )
    # Цикл 3: бюджет (DESCRIPTION, AMOUNT, CATEGORY)
    create_excel_file(
        out_dir / "project_report_budget_lines.xlsx",
        ("DESCRIPTION", "AMOUNT", "CATEGORY"),
        [
            ("Лицензии ПО", "120000", "Программное обеспечение"),
            ("Оборудование", "85000", "Инфраструктура"),
            ("Обучение", "45000", "Персонал"),
        ],
    )

    print("Готово. Файлы в каталоге tests/fixtures/")
    print("  Простые поля (формат имя/значение): *_values.xlsx")
    print("  Данные циклов: invoice_loop_items.xlsx, mixed_loop_works.xlsx, delivery_note_goods.xlsx,")
    print("                 project_report_participants.xlsx, project_report_milestones.xlsx, project_report_budget_lines.xlsx")
    print("  В приложении: вкладка цикла → «Загрузить из Excel...»")


if __name__ == "__main__":
    main()
